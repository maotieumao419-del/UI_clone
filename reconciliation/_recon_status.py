"""Join Order_Items NEW vs Sellerboard theo order_id, phân tích theo order_status."""
import os, re, sys, glob
from datetime import datetime
import openpyxl
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

IN = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "input")

# Dynamic File Matching
all_xlsx = glob.glob(os.path.join(IN, "*.xlsx"))
NEW = None
SB = None
for f in all_xlsx:
    fn = os.path.basename(f).upper()
    if fn.startswith("NEW_") or "SUMMARY_" in fn:
        NEW = f
    elif "DR_HAI_" in fn or "DASHBOARD_" in fn or "SELLERBOARD" in fn or "SB_" in fn:
        SB = f

if not NEW or not SB:
    print(f"Error: Could not match files dynamically in {IN}. NEW={NEW}, SB={SB}")
    sys.exit(1)

API_DB_COLUMNS = [
    'order_number', 'order_date', 'product', 'asin', 'sku', 'units', 'refunds', 'sales', 'promo', 'sellable_quota', 
    'refund_cost', 'amazon_fees', 'cost_of_goods', 'shipping', 'gross_profit', 'expenses', 'net_profit', 'margin', 
    'roi', 'coupon', 'row_type', 'updated_at', 'fee_state', 'order_status', 'price_source'
]

def num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, datetime):
        return round(v.day + v.month / 100, 2)
    try:
        return float(str(v).replace("$", "").replace(",", "").strip())
    except ValueError:
        return 0.0


def load(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    first_row = [str(c).strip() if c else "" for c in rows[0]]
    
    # Check if first row is actually data (headerless)
    is_headerless = False
    if re.search(r'\d{3}-\d{7}-\d{7}', first_row[0]):
        is_headerless = True
    else:
        first_row_norm = [c.replace("с", "c").lower() for c in first_row]
        if 'sku' not in first_row_norm and len(first_row) == len(API_DB_COLUMNS):
            is_headerless = True
            
    if is_headerless:
        print(f"Detected headerless Excel file: {path}. Applying schema columns.")
        return API_DB_COLUMNS, rows
    return first_row, rows[1:]


def col(hdr, name):
    nm = name.replace("с", "c").lower()
    for i, h in enumerate(hdr):
        if h.replace("с", "c").lower() == nm:
            return i
    return None


# NEW
nh, nr = load(NEW)
ni = {c: col(nh, c) for c in ("order_number", "sku", "sales", "amazon_fees", "order_status", "row_type")}
new = {}
for r in nr:
    if r[ni["row_type"]] != "normal":
        continue
    oid = re.search(r"\d{3}-\d{7}-\d{7}", str(r[ni["order_number"]]))
    oid = oid.group(0) if oid else str(r[ni["order_number"]])
    new[(oid, str(r[ni["sku"]]))] = {"sales": num(r[ni["sales"]]), "fees": num(r[ni["amazon_fees"]]),
                                     "status": r[ni["order_status"]]}

# SB
sh, sr = load(SB)
si = {c: col(sh, c) for c in ("Order number", "SKU", "Sales", "Amazon fees")}
sb = {}
sb_status = {}
for r in sr:
    raw = str(r[si["Order number"]])
    oid = re.search(r"\d{3}-\d{7}-\d{7}", raw)
    oid = oid.group(0) if oid else raw
    stt = raw.split(" / ")[1].strip() if " / " in raw else ""
    sb[(oid, str(r[si["SKU"]]))] = {"sales": num(r[si["Sales"]]), "fees": num(r[si["Amazon fees"]])}
    sb_status[oid] = stt

keys = set(new) | set(sb)
print(f"NEW rows={len(new)}  SB rows={len(sb)}  union keys={len(keys)}\n")

# Theo status của NEW
print("── Theo order_status (phía NEW) ──")
from collections import defaultdict
agg = defaultdict(lambda: {"n": 0, "ns": 0.0, "nf": 0.0, "ss": 0.0, "sf": 0.0, "in_sb": 0})
for k in keys:
    n = new.get(k); s = sb.get(k)
    st = n["status"] if n else "(chỉ có ở SB)"
    a = agg[st]
    a["n"] += 1
    if n:
        a["ns"] += n["sales"]; a["nf"] += n["fees"]
    if s:
        a["ss"] += s["sales"]; a["sf"] += s["fees"]; a["in_sb"] += 1
print(f"{'status':<16}{'#dòng':>6}{'NEW sales':>11}{'SB sales':>11}{'NEW fees':>10}{'SB fees':>10}{'#có ở SB':>9}")
for st, a in sorted(agg.items()):
    print(f"{str(st):<16}{a['n']:>6}{a['ns']:>11.2f}{a['ss']:>11.2f}{a['nf']:>10.2f}{a['sf']:>10.2f}{a['in_sb']:>9}")

# SB status distribution
sbcnt = defaultdict(int)
for v in sb_status.values():
    sbcnt[v] += 1
print("\n── Phân bố status phía Sellerboard ──")
for k, v in sorted(sbcnt.items()):
    print(f"  {k}: {v}")

# Tổng
print("\n── TỔNG ──")
print(f"  NEW: sales={sum(v['sales'] for v in new.values()):.2f}  fees={sum(v['fees'] for v in new.values()):.2f}")
print(f"  SB : sales={sum(v['sales'] for v in sb.values()):.2f}  fees={sum(v['fees'] for v in sb.values()):.2f}")
