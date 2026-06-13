"""So khớp toàn bộ cột 10/06: hệ thống (New_order_items, KHÔNG header, đọc theo vị trí)
vs Sellerboard (Dr_Hai_Craft, có header). Tách theo fee_state/order_status."""
import sys, re
from datetime import datetime
from collections import defaultdict
import openpyxl
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

IN = "data/input/"
F_SYS = IN + "New_order_items-10_06_2026 (1).xlsx"
F_SB = IN + "Dr_Hai_Craft_Dashboard_Order_Items_10_06_2026-10_06_2026_(2026_06_11_20_28_53_012).xlsx"

# vị trí cột file hệ thống (không header)
SYS = dict(order=0, sku=4, units=5, refunds=6, sales=7, promo=8, refund_cost=10,
           amazon_fees=11, cost_of_goods=12, shipping=13, gross_profit=14,
           net_profit=16, margin=17, roi=18, row_type=20, fee_state=22,
           order_status=23, price_source=24)
_CYR = str.maketrans({"с": "c", "о": "o", "а": "a", "е": "e", "р": "p"})
METRICS = ["units", "refunds", "sales", "promo", "refund_cost", "amazon_fees",
           "cost_of_goods", "shipping", "gross_profit", "net_profit", "margin", "roi"]


def num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, datetime):
        return round(v.day + v.month / 100, 2)
    try:
        return float(str(v).replace("$", "").replace(",", "").strip())
    except ValueError:
        return 0.0


def oid(v):
    m = re.search(r"\d{3}-\d{7}-\d{7}", str(v))
    return m.group(0) if m else str(v).split(" / ")[0].strip()


# ── hệ thống ──
wb = openpyxl.load_workbook(F_SYS, read_only=True, data_only=True)
sysrows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
A = defaultdict(lambda: {m: 0.0 for m in METRICS})
A_meta = {}
for r in sysrows:
    if r[SYS["row_type"]] != "normal":
        continue
    key = (oid(r[SYS["order"]]), str(r[SYS["sku"]]))
    for m in METRICS:
        A[key][m] += num(r[SYS[m]])
    A_meta[key] = (r[SYS["order_status"]], r[SYS["fee_state"]], r[SYS["price_source"]])

# ── Sellerboard ──
wb2 = openpyxl.load_workbook(F_SB, read_only=True, data_only=True)
sbrows = list(wb2[wb2.sheetnames[0]].iter_rows(values_only=True))
hdr = [str(c).translate(_CYR).strip().lower() for c in sbrows[0]]
SBcol = {
    "order": hdr.index("order number"), "sku": hdr.index("sku"),
    "units": hdr.index("units"), "refunds": hdr.index("refunds"),
    "sales": hdr.index("sales"), "promo": hdr.index("promo"),
    "refund_cost": hdr.index("refund cost"), "amazon_fees": hdr.index("amazon fees"),
    "cost_of_goods": hdr.index("cost of goods"), "shipping": hdr.index("shipping"),
    "gross_profit": hdr.index("gross profit"), "net_profit": hdr.index("net profit"),
    "margin": hdr.index("margin"), "roi": hdr.index("roi"),
}
B = defaultdict(lambda: {m: 0.0 for m in METRICS})
B_status = {}
for r in sbrows[1:]:
    if not r[SBcol["order"]]:
        continue
    key = (oid(r[SBcol["order"]]), str(r[SBcol["sku"]]))
    for m in METRICS:
        B[key][m] += num(r[SBcol[m]])
    mt = re.search(r"/\s*(\w+)\s*/", str(r[SBcol["order"]]))  # Shipped/Unshipped trong order#
    B_status[key] = mt.group(1) if mt else "?"

keys = set(A) | set(B)
print(f"Hệ thống: {len(A)} dòng normal | Sellerboard: {len(B)} dòng | tổng key: {len(keys)}\n")
print(f"{'CỘT':<15}{'HỆ THỐNG':>12}{'SELLERBOARD':>14}{'DELTA':>11}  #lệch")
print("-" * 60)
for m in METRICS:
    ta = round(sum(A.get(k, {}).get(m, 0) for k in keys), 2)
    tb = round(sum(B.get(k, {}).get(m, 0) for k in keys), 2)
    nbad = sum(1 for k in keys if abs(A.get(k, {}).get(m, 0) - B.get(k, {}).get(m, 0)) > 0.01)
    flag = "✅" if abs(ta - tb) < 0.5 and nbad == 0 else ("≈" if abs(ta - tb) < 0.5 else "❌")
    print(f"{m:<15}{ta:>12}{tb:>14}{ta-tb:>+11.2f}  {nbad:>3} {flag}")

# Rows lệch / chỉ 1 bên
only_sys = [k for k in A if k not in B]
only_sb = [k for k in B if k not in A]
print(f"\nChỉ có ở HỆ THỐNG ({len(only_sys)}): {only_sys[:6]}")
print(f"Chỉ có ở SELLERBOARD ({len(only_sb)}): {only_sb[:6]}")

# Phân tích amazon_fees theo fee_state / status (estimator lệch ở đâu)
print("\n── Amazon fees: lệch theo nhóm (chỉ key có ở cả 2) ──")
grp = defaultdict(lambda: [0.0, 0.0, 0])
for k in keys:
    if k in A and k in B:
        st, fs, ps = A_meta.get(k, ("?", "?", "?"))
        g = grp[(st, fs)]
        g[0] += A[k]["amazon_fees"]; g[1] += B[k]["amazon_fees"]; g[2] += 1
print(f"{'status / fee_state':<24}{'HỆ THỐNG':>11}{'SB':>10}{'DELTA':>10}  #")
for (st, fs), v in sorted(grp.items()):
    print(f"{str(st)+' / '+str(fs):<24}{v[0]:>11.2f}{v[1]:>10.2f}{v[0]-v[1]:>+10.2f}  {v[2]}")
