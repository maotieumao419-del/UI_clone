"""Mô-đun PPC đa-store: đọc nhiều file Excel quảng cáo, tính CTR/CVR, xuất CSV/Excel.

- Mỗi store = 1 file Excel (cấu hình ở settings.PPC_SOURCES).
- Sheet "Listing": danh sách SKU. 14+ sheet theo SKU: campaign + target + các khối
  [Impression, Click, Order, Note] lặp theo từng kỳ.
- Parser CHỊU LỖI: tìm cột theo từ khoá; tên campaign trống thì kế thừa dòng trên,
  rồi gom theo "Nhóm", cuối cùng gom vào "(Tất cả target)".
- CTR = Click/Impression; CVR = Order/Click (tính sẵn cho campaign & target).
"""
import csv
import io
import os
import threading
import time

import openpyxl

from ..config import settings
from . import vst_client

_RESERVED = {"listing", "portfolio id", "portfolio_id"}
_cache: dict[str, dict] = {}      # key=store -> {mtime, data}
_lock = threading.Lock()


# ---------------------------------------------------------------- helpers
def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace("\r", " ").strip()
    while "  " in s:
        s = s.replace("  ", " ")
    if s.lower().startswith("full sku:"):
        s = s[len("full sku:"):].strip()
    return s


def _to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", "").strip()))
    except (ValueError, TypeError):
        return None


def _norm_key(s: str) -> str:
    return _clean(s).upper().replace(" ", "")


def _pct(num, den):
    """Tỷ lệ phần trăm, làm tròn 2 chữ số; None nếu mẫu số = 0."""
    if not den:
        return None
    return round((num or 0) / den * 100, 2)


def _find_col(header: list[str], *keywords, exact=False) -> int | None:
    for i, h in enumerate(header):
        hl = _clean(h).lower()
        for kw in keywords:
            if (hl == kw) if exact else (kw in hl):
                return i
    return None


# ---------------------------------------------------------------- registry
def _ppc_dir() -> str:
    d = settings.PPC_DIR
    os.makedirs(d, exist_ok=True)
    return d


# -------------------------------------------------- nguồn API (DATA_SOURCE=vst)
_vst_fetched: dict[str, float] = {}     # key=store -> thời điểm kéo file gần nhất (epoch)


def _hydrate_from_vst(store: str | None) -> None:
    """Kéo file Excel PPC mới nhất từ API VST về PPC_DIR để parser cũ dùng lại.

    Có TTL (settings.VST_CACHE_TTL) để không gọi API mỗi request. Khi DATA_SOURCE=vst,
    file API ghi vào PPC_DIR cùng chỗ với file upload thủ công -> chạy song song được.
    """
    key = store or "__default__"
    ttl = settings.VST_CACHE_TTL
    if ttl > 0 and (time.time() - _vst_fetched.get(key, 0)) < ttl:
        return                                   # còn trong TTL, dùng file đã tải
    blob = vst_client.fetch_ppc_workbook_bytes(store)
    safe = (store or "vst_store").replace("\\", "_").replace("/", "_")
    with open(os.path.join(_ppc_dir(), f"{safe}.xlsx"), "wb") as f:
        f.write(blob)
    _vst_fetched[key] = time.time()
    _cache.pop(safe, None)                        # buộc parser đọc lại file mới


def _registry() -> dict[str, str]:
    """{store_label: filepath}. Nguồn = thư mục PPC_DIR (upload) + PPC_SOURCES cố định.

    Mỗi file .xlsx = 1 store. Nhãn = tên store khai báo, hoặc tên file (bỏ đuôi).
    Nhãn trùng được thêm hậu tố để không đè nhau.
    """
    reg: dict[str, str] = {}
    seen_paths: set[str] = set()

    def _add(label: str, path: str):
        ap = os.path.abspath(path)
        if ap in seen_paths:
            return
        base, n = label, 2
        while label in reg:
            label = f"{base} ({n})"
            n += 1
        reg[label] = path
        seen_paths.add(ap)

    # 1) Nguồn cố định (local dev) — bỏ qua file không tồn tại
    for src in settings.PPC_SOURCES:
        path = src.get("file", "")
        if path and os.path.exists(path):
            _add(src.get("store") or os.path.splitext(os.path.basename(path))[0], path)

    # 2) Quét thư mục upload trên server
    d = _ppc_dir()
    for fn in sorted(os.listdir(d)):
        if fn.lower().endswith((".xlsx", ".xlsm")) and not fn.startswith("~$"):
            _add(os.path.splitext(fn)[0], os.path.join(d, fn))

    return reg


def save_upload(filename: str, content: bytes) -> dict:
    """Lưu file Excel người dùng upload vào PPC_DIR -> trở thành 1 store mới."""
    safe = os.path.basename(filename).replace("\\", "_")
    if not safe.lower().endswith((".xlsx", ".xlsm")):
        return {"error": "Chỉ chấp nhận file .xlsx hoặc .xlsm"}
    path = os.path.join(_ppc_dir(), safe)
    with open(path, "wb") as f:
        f.write(content)
    _cache.pop(os.path.splitext(safe)[0], None)  # xoá cache nếu ghi đè
    return {"ok": True, "store": os.path.splitext(safe)[0], "file": safe, "stores": list_stores()}


def list_stores() -> list[dict]:
    return [{"store": s, "file": os.path.basename(p)} for s, p in _registry().items()]


def _default_store() -> str | None:
    reg = _registry()
    return next(iter(reg), None)


# ---------------------------------------------------------------- parsing
def _parse_listing(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [_clean(c) for c in rows[0]]
    c_sku = _find_col(header, "sku")
    c_store = _find_col(header, "store")
    c_link = _find_col(header, "link")
    c_pid = _find_col(header, "portfolio id", "portfolio_id")
    c_pname = _find_col(header, "portfolio name")
    c_status = _find_col(header, "status", "trạng thái")

    def g(r, i):
        return _clean(r[i]) if (i is not None and i < len(r)) else ""

    out, stt = [], 0
    for r in rows[1:]:
        sku = g(r, c_sku)
        if not sku:
            continue
        stt += 1
        out.append({
            "stt": stt, "sku": sku, "store": g(r, c_store), "link": g(r, c_link),
            "portfolio_id": g(r, c_pid), "portfolio_name": g(r, c_pname), "status": g(r, c_status),
        })
    return out


def _detect_metric_blocks(header: list[str], label_row: list[str]):
    blocks, i, n = [], 0, len(header)
    while i < n:
        if _clean(header[i]).lower().startswith("impression"):
            label = _clean(label_row[i]) if i < len(label_row) else ""
            if not label:
                label = f"Kỳ {len(blocks) + 1}"
            blk = {"period": label, "imp": i, "click": None, "order": None, "note": None}
            if i + 1 < n and "click" in _clean(header[i + 1]).lower():
                blk["click"] = i + 1
            if i + 2 < n and "order" in _clean(header[i + 2]).lower():
                blk["order"] = i + 2
            if i + 3 < n and _clean(header[i + 3]).lower() in ("note", "ghi chú"):
                blk["note"] = i + 3
            blocks.append(blk)
            i += 4
        else:
            i += 1
    return blocks


def _parse_sku_sheet(ws) -> dict:
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    header_idx = None
    for idx, row in enumerate(grid[:8]):
        if any(_clean(c).lower() == "campaign name" for c in row):
            header_idx = idx
            break
    if header_idx is None:
        for idx, row in enumerate(grid[:8]):
            if any(_clean(c).lower() == "target" for c in row):
                header_idx = idx
                break
    if header_idx is None:
        return {"periods": [], "campaigns": []}

    header = [_clean(c) for c in grid[header_idx]]
    label_row = grid[header_idx - 1] if header_idx > 0 else []
    c_camp = _find_col(header, "campaign name")
    c_type = _find_col(header, "loại campaign", "loai campaign")
    c_target = _find_col(header, "target")
    c_status = _find_col(header, "trạng thái", "trang thai", "status")
    c_group = _find_col(header, "nhóm", "nhom")
    blocks = _detect_metric_blocks(header, label_row)

    first_metric = blocks[0]["imp"] if blocks else len(header)
    c_tnote = None
    for i, h in enumerate(header):
        if i < first_metric and _clean(h).lower() in ("ghi chú", "ghi chu", "note"):
            c_tnote = i
            break

    def cell(row, idx):
        return row[idx] if (idx is not None and idx < len(row)) else None

    campaigns: dict[str, dict] = {}
    order_keys: list[str] = []
    current_name = None

    for row in grid[header_idx + 1:]:
        if not any(_clean(c) for c in row):
            continue
        raw_camp = _clean(cell(row, c_camp)) if c_camp is not None else ""
        if raw_camp:
            current_name = raw_camp
        name = current_name or (_clean(cell(row, c_group)) if c_group is not None else "") or "(Tất cả target)"
        target = _clean(cell(row, c_target))
        if not target and not raw_camp:
            continue

        if name not in campaigns:
            campaigns[name] = {
                "name": name, "type": "", "status": "",
                "totals": {"impression": 0, "click": 0, "order": 0},
                "targets": [],
            }
            order_keys.append(name)
        camp = campaigns[name]
        if not camp["type"]:
            camp["type"] = _clean(cell(row, c_type))
        if not camp["status"]:
            camp["status"] = _clean(cell(row, c_status))

        metrics = []
        t_imp = t_clk = t_odr = 0
        for blk in blocks:
            imp, clk, odr = _to_int(cell(row, blk["imp"])), _to_int(cell(row, blk["click"])), _to_int(cell(row, blk["order"]))
            metrics.append({"period": blk["period"], "impression": imp, "click": clk, "order": odr,
                            "note": _clean(cell(row, blk["note"]))})
            t_imp += imp or 0; t_clk += clk or 0; t_odr += odr or 0
            camp["totals"]["impression"] += imp or 0
            camp["totals"]["click"] += clk or 0
            camp["totals"]["order"] += odr or 0

        if target or metrics:
            camp["targets"].append({
                "target": target, "note": _clean(cell(row, c_tnote)), "metrics": metrics,
                "totals": {"impression": t_imp, "click": t_clk, "order": t_odr,
                           "ctr": _pct(t_clk, t_imp), "cvr": _pct(t_odr, t_clk)},
            })

    # CTR/CVR cấp campaign
    camps = []
    for k in order_keys:
        c = campaigns[k]
        tt = c["totals"]
        c["totals"]["ctr"] = _pct(tt["click"], tt["impression"])
        c["totals"]["cvr"] = _pct(tt["order"], tt["click"])
        camps.append(c)

    return {"periods": [b["period"] for b in blocks], "campaigns": camps}


# ---------------------------------------------------------------- cache/build
def _build(store: str, path: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    listing = _parse_listing(wb["Listing"]) if "Listing" in names else []

    sheet_index: dict[str, str] = {}
    for name in names:
        if name.strip().lower() in _RESERVED:
            continue
        first = None
        for r in wb[name].iter_rows(min_row=1, max_row=1, values_only=True):
            first = r[0] if r else None
            break
        for key in {_norm_key(name), _norm_key(first)}:
            if key:
                sheet_index.setdefault(key, name)

    for item in listing:
        item["has_detail"] = _norm_key(item["sku"]) in sheet_index

    return {"store": store, "file": os.path.basename(path), "path": path,
            "listing": listing, "sheet_index": sheet_index}


def _load(store: str | None) -> dict:
    # Nguồn API: kéo file mới nhất từ VST về (có TTL) rồi để logic file cũ chạy tiếp.
    if settings.DATA_SOURCE.lower() == "vst":
        try:
            _hydrate_from_vst(store)
        except vst_client.VSTError as e:
            return {"error": str(e), "listing": [], "sheet_index": {}, "store": store or ""}
    reg = _registry()
    if not reg:
        return {"error": "Chưa cấu hình file PPC nào (settings.PPC_SOURCES).",
                "listing": [], "sheet_index": {}, "store": store or ""}
    if store not in reg:
        store = _default_store()
    path = reg[store]
    mtime = os.path.getmtime(path)
    with _lock:
        c = _cache.get(store)
        if c is None or c["mtime"] != mtime:
            _cache[store] = {"mtime": mtime, "data": _build(store, path)}
        return _cache[store]["data"]


# ---------------------------------------------------------------- public API
def get_listing(store: str | None = None) -> dict:
    data = _load(store)
    return {
        "store": data.get("store"), "file": data.get("file"), "error": data.get("error"),
        "stores": list_stores(),
        "listing": [{k: v for k, v in it.items() if k != "_sheet"} for it in data.get("listing", [])],
    }


def get_sku_detail(sku: str, store: str | None = None) -> dict:
    data = _load(store)
    if data.get("error"):
        return {"error": data["error"], "sku": sku, "campaigns": [], "periods": []}
    sheet = data["sheet_index"].get(_norm_key(sku))
    if not sheet:
        return {"sku": sku, "store": data["store"], "sheet": None, "campaigns": [], "periods": [],
                "message": "SKU này chưa có sheet chi tiết campaign."}
    wb = openpyxl.load_workbook(data["path"], read_only=True, data_only=True)
    detail = _parse_sku_sheet(wb[sheet])
    detail.update({"sku": sku, "store": data["store"], "sheet": sheet})
    return detail


# ---------------------------------------------------------------- export
_EXPORT_HEADERS = ["Store", "SKU", "Sheet", "Campaign", "Loại", "Trạng thái", "Target",
                   "Note", "Kỳ", "Impression", "Click", "Order", "CTR(%)", "CVR(%)"]


def _flatten(store: str | None, sku: str | None) -> list[list]:
    """Bảng phẳng 1 dòng = 1 (target, kỳ) — dùng cho xuất CSV/Excel."""
    data = _load(store)
    store_label = data.get("store", "")
    skus = [sku] if sku else [it["sku"] for it in data.get("listing", []) if it.get("has_detail")]
    rows = []
    for sk in skus:
        detail = get_sku_detail(sk, store_label)
        sheet = detail.get("sheet") or ""
        for camp in detail.get("campaigns", []):
            for tg in camp.get("targets", []):
                ms = tg.get("metrics") or [{"period": "", "impression": None, "click": None, "order": None}]
                for m in ms:
                    rows.append([
                        store_label, sk, sheet, camp["name"], camp.get("type", ""), camp.get("status", ""),
                        tg.get("target", ""), tg.get("note", ""), m.get("period", ""),
                        m.get("impression"), m.get("click"), m.get("order"),
                        _pct(m.get("click"), m.get("impression")), _pct(m.get("order"), m.get("click")),
                    ])
    return rows


def export_csv(store: str | None, sku: str | None) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(_EXPORT_HEADERS)
    w.writerows(_flatten(store, sku))
    return ("﻿" + buf.getvalue()).encode("utf-8")  # BOM để Excel đọc đúng UTF-8


def export_xlsx(store: str | None, sku: str | None) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "PPC"
    ws.append(_EXPORT_HEADERS)
    for row in _flatten(store, sku):
        ws.append(row)
    # tô đậm hàng header
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
